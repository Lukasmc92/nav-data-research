import streamlit as st
from datetime import datetime, timedelta
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
import requests
from io import BytesIO

# --- Page config ---
st.set_page_config(page_title="NAV Data Pull", layout="wide")
st.title("📊 Closed-End Fund Data Research (Newly Optimized)")

# --- Download Tickers file ---
TICKERS_URL = "https://github.com/Lukasmc92/NAV-Tickers/raw/refs/heads/main/Tickers.xlsx"

@st.cache_data
def load_tickers():
    r = requests.get(TICKERS_URL)
    return pd.read_excel(BytesIO(r.content), engine="openpyxl")

df_tickers = load_tickers()
df_tickers = df_tickers.dropna(subset=["Fund", "NAV"])

fund_tickers = df_tickers["Fund"].tolist()
nav_tickers = df_tickers["NAV"].tolist()
fund_types = df_tickers["Fund Type"].tolist()
fund_subcats = df_tickers["Subcategory"].tolist()
fund_broadcats = df_tickers["Broad Category"].tolist()
fund_regions = df_tickers["Geographic Focus"].tolist()

# --- Date Picker ---
target_date = st.date_input(
    "Valuation Date (or last weekday before valuation date)",
    value=datetime.today(),
)
date_str = target_date.strftime('%Y-%m-%d')
start_date = (target_date - timedelta(days=2)).strftime('%Y-%m-%d')
end_date = (target_date + timedelta(days=2)).strftime('%Y-%m-%d')

# Helper function to get fundamentals as of a specific date
def get_fundamentals_asof_batch(ticker_list, as_of_date: str, quarterly=True):
    tickers = yf.Tickers(" ".join(ticker_list))
    as_of_date = pd.Timestamp(as_of_date)
    results = {}

    for ticker in ticker_list:
        t = tickers.tickers[ticker]
        balance = t.quarterly_balance_sheet if quarterly else t.balance_sheet

        if balance.empty:
            results[ticker] = {
                "shares_outstanding": None,
                "total_debt": None,
                "outside equity": None,
                "report_date": None
            }
            continue

        valid_dates = [d for d in balance.columns if d <= as_of_date]
        if not valid_dates:
            results[ticker] = {
                "shares_outstanding": None,
                "total_debt": None,
                "outside equity": None,
                "report_date": None
            }
            continue

        latest = max(valid_dates)

        shares = next((balance.loc[row, latest] for row in ["Ordinary Shares Number", "Share Issued"] if row in balance.index), None)
        debt = next((balance.loc[row, latest] for row in ["Total Debt", "Long Term Debt", "Current Debt"] if row in balance.index), None)
        otequity = next((balance.loc[row, latest] for row in ["Preferred Securities Outside Stock Equity"] if row in balance.index), None)

        results[ticker] = {
            "shares_outstanding": shares,
            "total_debt": debt,
            "outside equity": otequity,
            "report_date": latest
        }

    return results

# Helper function to retry missing prices
def fetch_missing_prices(tickers, date_str, start_date, end_date, close_prices):
    for ticker in tickers:
        if ticker not in close_prices.columns or pd.isna(close_prices.loc[date_str, ticker]):
            try:
                data = yf.download(ticker, start=start_date, end=end_date, progress=False)
                if not data.empty:
                    data.index = data.index.strftime("%Y-%m-%d")
                    if date_str in data.index:
                        close_prices.loc[date_str, ticker] = data.loc[date_str, "Close"]
            except Exception as e:
                print(f"Retry failed for {ticker}: {e}")
    return close_prices

# --- Run Button ---
if st.button("Download NAV Data"):

    # --- Download all prices in bulk ---
    st.info("⏳ Downloading price data in bulk from Yahoo Finance...")
    tickers_all = list(set(fund_tickers + nav_tickers))

    prices = yf.download(
        tickers_all,
        start=start_date,
        end=end_date,
        auto_adjust=False,
        group_by="ticker",
        progress=False
    )

    # --- Normalize columns: handle single vs multi ticker ---
    if isinstance(prices.columns, pd.MultiIndex):
        close_prices = prices.xs("Close", axis=1, level=1)
    else:
        close_prices = prices.to_frame(name=tickers_all[0]) if len(tickers_all) == 1 else prices

    close_prices.index = close_prices.index.strftime("%Y-%m-%d")

    # Fallback to nearest available date if date_str is missing
    if date_str not in close_prices.index:
        available_dates = close_prices.index.tolist()
        fallback_date = max([d for d in available_dates if d <= date_str], default=None)
        if fallback_date:
            st.warning(f"No data for {date_str}. Using fallback date: {fallback_date}")
            date_str = fallback_date
        else:
            st.error(f"No data available near {date_str}. Try another date.")
            st.stop()

    # Retry missing prices
    close_prices = fetch_missing_prices(tickers_all, date_str, start_date, end_date, close_prices)

    # --- Batch get fundamentals ---
    st.info("⏳ Downloading fundamentals (shares/debt)...")
    tickers_obj = yf.Tickers(fund_tickers)
    fundamentals_batch = get_fundamentals_asof_batch(fund_tickers, target_date)

    rows = []
    progress_bar = st.progress(0)

    for idx, (fund, nav, types, subcategories, broadcats, regions) in enumerate(
        zip(fund_tickers, nav_tickers, fund_types, fund_subcats, fund_broadcats, fund_regions)
    ):
        fund_price = close_prices.loc[date_str, fund] if fund in close_prices.columns else None
        nav_price = close_prices.loc[date_str, nav] if nav in close_prices.columns else None
        discount = (fund_price / nav_price) if (fund_price and nav_price) else None

        fundamentals = fundamentals_batch.get(fund, {})
        shares_outstanding = fundamentals.get("shares_outstanding")
        total_debt = fundamentals.get("total_debt")
        outside_equity = fundamentals.get("outside equity")

        shares_millions = round(shares_outstanding / 1_000_000, 2) if shares_outstanding else None
        debt_millions = round(total_debt / 1_000_000, 2) if total_debt else None
        outside_equity_millions = round(outside_equity / 1_000_000, 2) if outside_equity else None

        ticker_obj = yf.Ticker(fund)
        info = ticker_obj.info
        fund_name = info.get("longName", fund)

        rows.append([
            fund_name, broadcats, types, subcategories, regions, date_str,
            fund, fund_price, nav, nav_price, discount, shares_millions, debt_millions, outside_equity_millions
        ])

        progress_bar.progress((idx + 1) / len(fund_tickers))

    df = pd.DataFrame(rows, columns=[
        "Fund Name", "Broad Category", "Fund Type", "Subcategory", "Geographic Focus", "Date",
        "Fund Ticker", "Fund Close Price", "NAV Ticker", "NAV Close Price", "Discount",
        "Shares Outstanding(M)", "Total Debt(M)", "Outside Equity (M)"
    ])

    excel_filename = f'Closed_End_Fund_Data_{date_str}.xlsx'
    df.to_excel(excel_filename, index=False, sheet_name='Sheet1')

    wb = load_workbook(excel_filename)
    ws = wb['Sheet1']
    message_row = ws.max_row + 2
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    method = "This file was created using Python, Streamlit, and batched Yahoo Finance requests."
    ws.cell(row=message_row, column=1, value=f"Downloaded on {timestamp}. Method: {method}")
    wb.save(excel_filename)

    st.success("✅ NAV Data Pull Complete")
    st.dataframe(df)

    with open(excel_filename, "rb") as f:
        st.download_button(
            label="📥 Download Excel File",
            data=f,
            file_name=excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )




